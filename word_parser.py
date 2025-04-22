import os
from docx import Document
from openpyxl import Workbook, load_workbook
import re
from datetime import datetime

def check_file_format(file_path, file_type):
    """
    Проверяет формат файла и его существование
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл не найден: {file_path}")
    
    if file_type == 'word' and not file_path.lower().endswith('.docx'):
        raise ValueError(
            "Неверный формат файла Word. Пожалуйста, используйте файл в формате .docx\n"
            "Если у вас файл в формате .doc, пересохраните его в формате .docx через Microsoft Word"
        )
    elif file_type == 'excel' and not file_path.lower().endswith('.xlsx'):
        raise ValueError(
            "Неверный формат файла Excel. Пожалуйста, используйте файл в формате .xlsx"
        )

def safe_create_excel(excel_path):
    """
    Безопасно создает новый Excel файл
    """
    try:
        # Если файл существует и поврежден, удаляем его
        if os.path.exists(excel_path):
            try:
                load_workbook(excel_path)
            except:
                print(f"Обнаружен поврежденный файл Excel. Создаем новый файл: {excel_path}")
                os.remove(excel_path)
        
        # Если файл не существует или был удален, создаем новый
        if not os.path.exists(excel_path):
            wb = Workbook()
            ws = wb.active
            headers = ['№', 'Дата исследования', 'Исследование', 'ФИО', 'Возраст', 'Отделение', 'Заключение']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            wb.save(excel_path)
            return wb
        
        return load_workbook(excel_path)
    except Exception as e:
        raise Exception(f"Не удалось создать/открыть файл Excel: {str(e)}")

def clean_department(text):
    """
    Очищает текст отделения от лишней информации
    """
    # Удаляем информацию о палате, если она есть
    department = text.split(',')[0] if ',' in text else text
    # Удаляем "Палата:" и всё после этого, если такой текст есть
    department = department.split('Палата:')[0]
    return department.strip()

def extract_data_from_word(doc_path):
    """
    Извлекает данные из Word документа
    """
    try:
        check_file_format(doc_path, 'word')
        doc = Document(doc_path)
    except FileNotFoundError as e:
        raise e
    except ValueError as e:
        raise e
    except Exception as e:
        raise Exception(
            f"Ошибка при открытии файла {doc_path}.\n"
            f"Возможно файл поврежден или защищен паролем.\n"
            f"Попробуйте открыть файл в Microsoft Word и сохранить его заново в формате .docx"
        )

    all_data = []
    current_data = None
    paragraphs = list(doc.paragraphs)
    
    for i, paragraph in enumerate(paragraphs):
        text = paragraph.text.strip()
        if not text:
            continue
            
        # Начало нового описания ЭКГ
        if "ОПИСАНИЕ ЭЛЕКТРОКАРДИОГРАММЫ" in text:
            if current_data:
                # Очищаем отделение перед добавлением записи
                if current_data['department']:
                    current_data['department'] = clean_department(current_data['department'])
                all_data.append(current_data)
            current_data = {
                'date': '',
                'research': 'ЭКГ',
                'fio': '',
                'age': '',
                'department': '',
                'conclusion': ''
            }
            continue
            
        if not current_data:
            continue
            
        # Извлечение ФИО
        if text.startswith("ФИО:"):
            current_data['fio'] = text.replace("ФИО:", "").strip()
            
        # Извлечение возраста (даты рождения)
        elif text.startswith("Возраст:"):
            current_data['age'] = text.replace("Возраст:", "").strip()
            
        # Извлечение отделения
        elif text.startswith("Отделение:"):
            department_text = text.replace("Отделение:", "").strip()
            current_data['department'] = department_text
            
        # Извлечение даты
        elif text.startswith("Дата описания:"):
            date_match = re.search(r'(\d{2}\.\d{2}\.\d{4})', text)
            if date_match:
                current_data['date'] = date_match.group(1)
                
        # Извлечение заключения
        elif text.startswith("Заключение:"):
            conclusion = []
            # Добавляем первую строку заключения
            first_line = text.replace("Заключение:", "").strip()
            if first_line:
                conclusion.append(first_line)
            
            # Ищем следующие параграфы, которые могут быть частью заключения
            next_index = i + 1
            while next_index < len(paragraphs):
                next_text = paragraphs[next_index].text.strip()
                if not next_text or "ОПИСАНИЕ ЭЛЕКТРОКАРДИОГРАММЫ" in next_text:
                    break
                if not any(next_text.startswith(prefix) for prefix in ["ФИО:", "Возраст:", "Отделение:", "Палата:", "Дата описания:"]):
                    conclusion.append(next_text)
                else:
                    break
                next_index += 1
            
            current_data['conclusion'] = " ".join(conclusion)
    
    # Добавляем последнее описание
    if current_data:
        # Очищаем отделение перед добавлением последней записи
        if current_data['department']:
            current_data['department'] = clean_department(current_data['department'])
        all_data.append(current_data)
    
    return all_data

def write_to_excel(all_data, excel_path):
    """
    Записывает данные в Excel файл
    """
    try:
        check_file_format(excel_path, 'excel')
        wb = safe_create_excel(excel_path)
        ws = wb.active
        
        # Определяем начальный номер строки
        start_row = ws.max_row + 1 if ws.max_row > 1 else 2
        
        # Записываем все данные
        for index, data in enumerate(all_data, 0):
            row = start_row + index
            
            # Записываем номер строки
            ws.cell(row=row, column=1, value=row-1)
            
            # Записываем данные
            ws.cell(row=row, column=2, value=data['date'])
            ws.cell(row=row, column=3, value=data['research'])
            ws.cell(row=row, column=4, value=data['fio'])
            ws.cell(row=row, column=5, value=data['age'])
            ws.cell(row=row, column=6, value=data['department'])
            ws.cell(row=row, column=7, value=data['conclusion'])
        
        # Настраиваем ширину колонок
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Ограничиваем максимальную ширину
            ws.column_dimensions[column].width = adjusted_width
        
        # Сохраняем файл
        try:
            wb.save(excel_path)
        except PermissionError:
            raise Exception(
                f"Не удалось сохранить файл {excel_path}.\n"
                "Возможно, файл открыт в Excel. Пожалуйста, закройте его и попробуйте снова."
            )
    except Exception as e:
        if "zip file" in str(e):
            raise Exception(
                f"Ошибка при работе с файлом Excel {excel_path}.\n"
                "Файл поврежден или имеет неправильный формат.\n"
                "Программа попытается создать новый файл при следующем запуске."
            )
        raise Exception(f"Ошибка при сохранении в Excel: {str(e)}")

def process_word_document(word_path, excel_path):
    """
    Обрабатывает Word документ и записывает данные в Excel
    """
    try:
        print(f"Обработка файла Word: {word_path}")
        all_data = extract_data_from_word(word_path)
        
        if not all_data:
            print("Внимание: Не найдено ни одной записи в документе")
            return
            
        print(f"Найдено записей: {len(all_data)}")
        print(f"Сохранение данных в Excel файл: {excel_path}")
        write_to_excel(all_data, excel_path)
        print(f"Данные успешно записаны в файл: {excel_path}")
        print(f"Обработано записей: {len(all_data)}")
        
    except Exception as e:
        print(f"Произошла ошибка: {str(e)}")

if __name__ == "__main__":
    # Пример использования
    word_path = "input.docx"  # Путь к Word документу
    excel_path = "output.xlsx"  # Путь к Excel файлу
    process_word_document(word_path, excel_path) 